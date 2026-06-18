#!/usr/bin/env python3
"""
Extract values of a given IFC property (by property name) from all IFC files in a folder.

Example:
  python extract_ifc_property_values.py \
    --folder "C:/in/ifcs" \
    --modelFolder "C:/in/ifcs/models" \
    --propertyName "Reference" \
    --out "C:/out/values.txt"
"""

from __future__ import annotations

import argparse
import os
import sys
import re
from pathlib import Path
from typing import Any, Iterable, List, Optional, Dict, Tuple

import ifcopenshell
import openpyxl


def _iter_ifc_files(root: Path) -> Iterable[Path]:
    # Recursively find .ifc files (case-insensitive)
    for p in root.rglob("*"):
        if p.is_file() and p.suffix.lower() == ".ifc":
            yield p


def _stringify_ifc_value(nominal_value: Any) -> str:
    """
    Convert IFC nominal value objects / python primitives to a stable string.
    ifcopenshell often returns:
      - python primitives (str/int/float/bool)
      - IFC wrapper objects with .wrappedValue
    """
    if nominal_value is None:
        return ""
    if hasattr(nominal_value, "wrappedValue"):
        v = getattr(nominal_value, "wrappedValue")
        return "" if v is None else _propertyStringTreatment(v)
    return str(nominal_value)

def _propertyStringTreatment(p):
    p = p.replace("[L]%","")
    p = p[:-2]
    p = p.replace(".","")
    p = re.sub(r"\d","", p)
    return str(p)


def _extract_property_values_from_object(obj, pset_name: str, prop_name: str):
    values = []

    rels = getattr(obj, "IsDefinedBy", None)
    if not rels:
        return values

    for rel in rels:
        if not rel.is_a("IfcRelDefinesByProperties"):
            continue

        pdef = rel.RelatingPropertyDefinition
        if not pdef or not pdef.is_a("IfcPropertySet"):
            continue

        if str(pdef.Name) != pset_name:
            continue

        for prop in pdef.HasProperties or []:
            if str(prop.Name) != prop_name:
                continue

            if prop.is_a("IfcPropertySingleValue"):
                values.append(_stringify_ifc_value(prop.NominalValue))

            elif prop.is_a("IfcPropertyEnumeratedValue"):
                values.append(
                    ";".join(_stringify_ifc_value(v) for v in prop.EnumerationValues or [])
                )

            elif prop.is_a("IfcPropertyListValue"):
                values.append(
                    ";".join(_stringify_ifc_value(v) for v in prop.ListValues or [])
                )

    return values


def extract_values_from_ifc_file(ifc_path: Path, property_name: str) -> List[str]:
    try:
        model = ifcopenshell.open(str(ifc_path))
    except Exception as e:
        print(f"[WARN] Failed to open: {ifc_path} ({e})", file=sys.stderr)
        return []

    found: List[str] = []
    # By spec, most element-like objects are IfcObjectDefinition; this covers elements, types, spatial, etc.
    for obj in model.by_type("IfcObjectDefinition"):
        try:
            found.extend(_extract_property_values_from_object(obj, property_name.split(".")[0], property_name.split(".")[1]))
        except Exception as e:
            # Keep going; IFCs can contain oddities
            oid = getattr(obj, "GlobalId", None)
            print(f"[WARN] Error on object {oid} in {ifc_path.name}: {e}", file=sys.stderr)

    return found

def load_css_hovedbegreb_lookup(
    xlsx_path: str | Path,
    css_col: str = "CCS",
    hovedbegreb_col: str = "Hovedbegreb",
    normalize: bool = True,
) -> Dict[str, str]:
    """
    Reads an .xlsx containing two sheets, each with columns:
      - CCS
      - Hovedbegreb

    Returns a dict: {ccs_value -> hovedbegreb_value}
    If a CCS value appears multiple times across sheets, the first non-empty Hovedbegreb wins.
    """
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    lookup: Dict[str, str] = {}

    def norm(s: object) -> str:
        if s is None:
            return ""
        t = str(s)
        return t.strip() if normalize else t

    for ws in wb.worksheets:
        # Map header -> column index
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue

        headers = {norm(v): idx for idx, v in enumerate(header_row) if norm(v)}
        if css_col not in headers or hovedbegreb_col not in headers:
            continue

        css_idx = headers[css_col]
        hov_idx = headers[hovedbegreb_col]

        for row in ws.iter_rows(min_row=2, values_only=True):
            css_val = norm(row[css_idx] if css_idx < len(row) else None)
            hov_val = norm(row[hov_idx] if hov_idx < len(row) else None)
            if not css_val:
                continue

            # Keep first non-empty mapping
            if css_val not in lookup and hov_val:
                lookup[css_val] = hov_val
            elif css_val not in lookup and not hov_val:
                # store empty only if nothing else exists (optional behavior)
                lookup[css_val] = ""

    wb.close()
    return lookup


def match_and_concat(
    property_value: str,
    lookup: Dict[str, str],
    sep: str = " ",
    default_hovedbegreb: str = "",
) -> Tuple[str, Optional[str]]:
    """
    Returns:
      (concatenated_output, matched_hovedbegreb_or_None)

    Output is: "<property_value><sep><hovedbegreb>" if matched, otherwise "<property_value>".
    """
    key = property_value.strip()
    hovedbegreb = lookup.get(key)

    if hovedbegreb is None or hovedbegreb == "":
        if default_hovedbegreb:
            return f"{property_value}{sep}{default_hovedbegreb}", default_hovedbegreb
        return property_value, None

    return f"{property_value}{sep}{hovedbegreb}", hovedbegreb


def main() -> int:
    ap = argparse.ArgumentParser(description="Extract IFC property values from all IFC files in a folder.")
    ap.add_argument("--folder", required=True, help="Root folder containing IFC files (scanned recursively).")
    ap.add_argument("--modelFolder", required=True, help="Model folder argument (accepted as input; used as a subfolder if relative).")
    ap.add_argument("--propertyName", required=True, help="Property name to look for (exact match).")
    ap.add_argument("--out", default="property_values.txt", help="Output txt file path.")
    ap.add_argument("--dedupe", action="store_true", help="Remove duplicate values (preserves first-seen order).")
    ap.add_argument("--skip-empty", action="store_true", help="Skip empty-string values.")
    args = ap.parse_args()

    folder = Path(args.folder).expanduser().resolve()

    # Interpret modelFolder: if it's relative, treat it as relative to --folder; else use as given.
    model_folder = Path(args.modelFolder).expanduser()
    if not model_folder.is_absolute():
        model_folder = (folder / model_folder).resolve()
    else:
        model_folder = model_folder.resolve()

    # Decide where to scan:
    # - If modelFolder exists and is inside folder (or user wants that), scan modelFolder
    # - Else scan folder
    scan_root = model_folder if model_folder.exists() else folder

    if not scan_root.exists():
        print(f"[ERROR] Scan root does not exist: {scan_root}", file=sys.stderr)
        return 2

    out_path = Path(args.out).expanduser()
    if not out_path.is_absolute():
        out_path = (Path.cwd() / out_path).resolve()

    all_values: List[str] = []
    ifc_files = list(_iter_ifc_files(scan_root))
    if not ifc_files:
        print(f"[WARN] No IFC files found under: {scan_root}", file=sys.stderr)

    for ifc_path in ifc_files:
        vals = extract_values_from_ifc_file(ifc_path, args.propertyName)
        all_values.extend(vals)

    if args.skip_empty:
        all_values = [v for v in all_values if v != ""]

    if args.dedupe:
        seen = set()
        deduped: List[str] = []
        for v in all_values:
            if v in seen:
                continue
            seen.add(v)
            deduped.append(v)
        all_values = deduped

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="\n") as f:
        for v in sorted(all_values):
            lookup = load_css_hovedbegreb_lookup(r"CCS.xlsx")
            vOut, hov = match_and_concat(v, lookup)
            f.write(f"{vOut}\n")

    print(f"Wrote {len(all_values)} value(s) to: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
